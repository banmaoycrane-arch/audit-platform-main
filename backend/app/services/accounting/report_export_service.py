"""财务报表导出：xlsx / csv / zip 打包（标准列报格式）。"""

from __future__ import annotations



import csv

import io

import zipfile

from typing import Any



from openpyxl import Workbook



from app.services.accounting.report_format_standard import (

    append_balance_sheet_body,

    append_cash_flow_body,

    append_income_statement_body,

    append_trial_balance_body,

    report_meta_lines,

)



SUPPORTED_FORMATS = {"xlsx", "csv", "pdf"}





def _workbook_to_bytes(wb: Workbook) -> bytes:

    buf = io.BytesIO()

    wb.save(buf)

    return buf.getvalue()





def _cell_display_width(value: Any) -> float:
    text = str(value) if value is not None else ""
    width = 0.0
    for ch in text:
        width += 2.0 if ord(ch) > 127 else 1.0
    return width


def _autofit_worksheet(ws, rows: list[list[Any]], *, max_width: float = 42.0) -> None:
    from openpyxl.utils import get_column_letter

    if not rows:
        return
    col_count = max(len(row) for row in rows)
    widths = [8.0] * col_count
    for row in rows:
        for idx, cell in enumerate(row):
            widths[idx] = min(max(widths[idx], _cell_display_width(cell) + 2.0), max_width)
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def _format_classic_worksheet(ws, rows: list[list[Any]]) -> None:
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    if not rows:
        return
    col_count = max(len(row) for row in rows if row)
    if col_count <= 0:
        return

    title_row = rows[0] if len(rows[0]) == 1 else None
    if title_row:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
        cell = ws.cell(row=1, column=1)
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    meta_row_idx = 2 if title_row else None
    if meta_row_idx and meta_row_idx <= len(rows):
        meta = rows[meta_row_idx - 1]
        if len(meta) >= 3:
            span = max(col_count // 3, 1)
            ws.merge_cells(start_row=meta_row_idx, start_column=1, end_row=meta_row_idx, end_column=span)
            ws.merge_cells(
                start_row=meta_row_idx,
                start_column=span + 1,
                end_row=meta_row_idx,
                end_column=min(span * 2, col_count),
            )
            if col_count > span * 2:
                ws.merge_cells(
                    start_row=meta_row_idx,
                    start_column=span * 2 + 1,
                    end_row=meta_row_idx,
                    end_column=col_count,
                )
            for col in range(1, col_count + 1):
                ws.cell(row=meta_row_idx, column=col).alignment = Alignment(
                    horizontal="left" if col == 1 else ("right" if col > span * 2 else "center"),
                    vertical="center",
                    wrap_text=True,
                )

    for row_idx, row in enumerate(rows, start=1):
        if not row:
            continue
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(value, (int, float)):
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_idx > 1 and row_idx > 3 and str(value).replace(".", "", 1).replace("-", "", 1).isdigit():
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(vertical="center", wrap_text=col_idx <= 2)

    _autofit_worksheet(ws, rows)


def _write_rows_to_sheet(ws, rows: list[list[Any]], *, classic_format: bool = False) -> None:
    for row in rows:
        ws.append(row)
    if classic_format:
        _format_classic_worksheet(ws, rows)
    else:
        _autofit_worksheet(ws, rows)





def _report_rows_with_meta(
    report_title: str,
    report: dict[str, Any],
    body_builder,
    *,
    ledger_name: str | None = None,
    as_of_date: Any = None,
    report_kind: str | None = None,
) -> list[list[Any]]:
    from app.services.accounting.classic_report_layout_service import (
        classic_report_footer_rows,
        classic_report_header_rows,
    )

    if report.get("format") and report_kind:
        rows: list[list[Any]] = classic_report_header_rows(report_kind, report)
    else:
        rows = report_meta_lines(
            report_title=report_title,
            ledger_name=ledger_name or report.get("ledger_name"),
            period_code=report.get("period_code"),
            as_of_date=as_of_date or report.get("as_of_date"),
        )
    body_builder(rows, report)
    if report.get("format") and report_kind:
        rows.extend(classic_report_footer_rows(report))
    return rows





def trial_balance_to_xlsx(report: dict[str, Any], *, ledger_name: str | None = None) -> bytes:

    wb = Workbook()

    ws = wb.active

    ws.title = "科目余额表"

    rows = _report_rows_with_meta("科目余额表", report, append_trial_balance_body, ledger_name=ledger_name)

    _write_rows_to_sheet(ws, rows)

    return _workbook_to_bytes(wb)





def trial_balance_to_csv(report: dict[str, Any], *, ledger_name: str | None = None) -> bytes:

    buf = io.StringIO()

    writer = csv.writer(buf)

    rows = _report_rows_with_meta("科目余额表", report, append_trial_balance_body, ledger_name=ledger_name)

    for row in rows:

        writer.writerow(row)

    return ("\ufeff" + buf.getvalue()).encode("utf-8")





def balance_sheet_to_xlsx(report: dict[str, Any], *, ledger_name: str | None = None) -> bytes:

    wb = Workbook()

    ws = wb.active

    ws.title = "资产负债表"

    rows = _report_rows_with_meta("资产负债表", report, append_balance_sheet_body, ledger_name=ledger_name, report_kind="balance_sheet")

    _write_rows_to_sheet(ws, rows, classic_format=str(report.get("format", "")).startswith("classic"))

    return _workbook_to_bytes(wb)





def balance_sheet_to_csv(report: dict[str, Any], *, ledger_name: str | None = None) -> bytes:

    buf = io.StringIO()

    writer = csv.writer(buf)

    rows = _report_rows_with_meta("资产负债表", report, append_balance_sheet_body, ledger_name=ledger_name, report_kind="balance_sheet")

    for row in rows:

        writer.writerow(row)

    return ("\ufeff" + buf.getvalue()).encode("utf-8")





def income_statement_to_xlsx(report: dict[str, Any], *, ledger_name: str | None = None) -> bytes:

    wb = Workbook()

    ws = wb.active

    ws.title = "损益表"

    rows = _report_rows_with_meta("损益表", report, append_income_statement_body, ledger_name=ledger_name, report_kind="income_statement")

    _write_rows_to_sheet(ws, rows, classic_format=str(report.get("format", "")).startswith("classic"))

    return _workbook_to_bytes(wb)





def income_statement_to_csv(report: dict[str, Any], *, ledger_name: str | None = None) -> bytes:

    buf = io.StringIO()

    writer = csv.writer(buf)

    rows = _report_rows_with_meta("损益表", report, append_income_statement_body, ledger_name=ledger_name, report_kind="income_statement")

    for row in rows:

        writer.writerow(row)

    return ("\ufeff" + buf.getvalue()).encode("utf-8")





def cash_flow_to_xlsx(report: dict[str, Any], *, ledger_name: str | None = None) -> bytes:

    wb = Workbook()

    ws = wb.active

    ws.title = "现金流量表"

    rows = _report_rows_with_meta("现金流量表", report, append_cash_flow_body, ledger_name=ledger_name, report_kind="cash_flow")

    _write_rows_to_sheet(ws, rows, classic_format=str(report.get("format", "")).startswith("classic"))

    return _workbook_to_bytes(wb)





def cash_flow_to_csv(report: dict[str, Any], *, ledger_name: str | None = None) -> bytes:

    buf = io.StringIO()

    writer = csv.writer(buf)

    rows = _report_rows_with_meta("现金流量表", report, append_cash_flow_body, ledger_name=ledger_name, report_kind="cash_flow")

    for row in rows:

        writer.writerow(row)

    return ("\ufeff" + buf.getvalue()).encode("utf-8")





def build_reports_zip_package(

    files: list[tuple[str, bytes]],

) -> bytes:

    """将多个报表文件打包为 zip。"""

    buffer = io.BytesIO()

    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as zf:

        for filename, content in files:

            zf.writestr(filename, content)

    return buffer.getvalue()





EXPORT_BUILDERS = {

    "trial_balance": {"xlsx": trial_balance_to_xlsx, "csv": trial_balance_to_csv},

    "balance_sheet": {"xlsx": balance_sheet_to_xlsx, "csv": balance_sheet_to_csv},

    "income_statement": {"xlsx": income_statement_to_xlsx, "csv": income_statement_to_csv},

    "cash_flow": {"xlsx": cash_flow_to_xlsx, "csv": cash_flow_to_csv},

}



CONTENT_TYPES = {

    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",

    "csv": "text/csv; charset=utf-8",

    "pdf": "application/pdf",

    "zip": "application/zip",

}



REPORT_XLSX_FILENAMES = {

    "trial_balance": "01_科目余额表.xlsx",

    "balance_sheet": "02_资产负债表.xlsx",

    "income_statement": "03_损益表.xlsx",

    "cash_flow": "04_现金流量表.xlsx",

}



REPORT_PDF_FILENAMES = {

    "trial_balance": "01_科目余额表_签章版.pdf",

    "balance_sheet": "02_资产负债表_签章版.pdf",

    "income_statement": "03_损益表_签章版.pdf",

    "cash_flow": "04_现金流量表_签章版.pdf",

}



SUBSIDIARY_XLSX_FILENAME = "05_明细账.xlsx"


