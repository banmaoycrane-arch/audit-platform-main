"""Excel 序时簿蓝字/红字金额解析。

会计口径：
- 蓝字：在该行所属借贷列上记正数发生额（借方列增加借方，贷方列增加贷方）
- 红字：在借方列或贷方列均可能出现，在同侧记负数冲减（不是到对方科目再记蓝字）
- 一张凭证内蓝字/红字合计后仍须借贷相等，才能作为完整事务保存
"""

from __future__ import annotations

from decimal import Decimal
from pathlib import Path
from typing import Any, Literal

import pandas as pd
from openpyxl import load_workbook

from app.money import parse_decimal
from app.services.doc_parsing.format_template import match_header

ColorHint = Literal["blue", "red", "default"]
AmountSide = Literal["debit", "credit", "unknown"]

_AMOUNT_HEADER_KEYWORDS = ("借", "贷", "debit", "credit", "发生额", "amount", "金额", "本位币")
_REVERSAL_SUMMARY_KEYWORDS = ("冲", "红字", "更正", "冲销", "冲账", "作废", "错账")


def _amount(value: Any) -> Decimal:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return Decimal("0.00")
    try:
        cleaned = str(value).replace(",", "").replace("，", "").replace("¥", "").replace("￥", "").strip()
        return parse_decimal(cleaned, decimal_places=2, allow_empty=True)
    except Exception:
        return Decimal("0.00")


def is_reversal_summary(summary: str) -> bool:
    text = (summary or "").strip()
    return any(keyword in text for keyword in _REVERSAL_SUMMARY_KEYWORDS)


def classify_font_color(cell: Any) -> ColorHint:
    """根据 Excel 单元格字体颜色辅助识别蓝字/红字（部分导出软件会用颜色标记）。"""
    if cell is None:
        return "default"
    font = getattr(cell, "font", None)
    if font is None or font.color is None:
        return "default"

    color = font.color
    rgb: str | None = None
    if getattr(color, "type", None) == "rgb" and color.rgb:
        rgb = str(color.rgb)
    elif getattr(color, "type", None) == "indexed" and color.indexed is not None:
        indexed_map = {
            10: "FF0000",
            2: "FF0000",
            12: "0000FF",
            4: "0000FF",
            5: "0000FF",
        }
        rgb = indexed_map.get(int(color.indexed))

    if not rgb:
        return "default"

    hex_text = rgb[-6:].upper()
    if len(hex_text) != 6:
        return "default"
    try:
        red = int(hex_text[0:2], 16)
        green = int(hex_text[2:4], 16)
        blue = int(hex_text[4:6], 16)
    except ValueError:
        return "default"

    if red >= 150 and green <= 120 and blue <= 120:
        return "red"
    if blue >= 120 and red <= 120 and green <= 180:
        return "blue"
    if red >= 180 and blue >= 120 and green <= 100:
        return "red"
    return "default"


def is_amount_column(header: str, field_name: str | None = None) -> bool:
    if field_name in {"debit_amount", "credit_amount"}:
        return True
    text = (header or "").strip().lower()
    return any(keyword in text or keyword in (header or "") for keyword in _AMOUNT_HEADER_KEYWORDS)


def column_amount_side(header: str, field_name: str | None = None) -> AmountSide:
    header_text = str(header or "").strip()
    if field_name == "debit_amount" or ("借" in header_text and "贷" not in header_text):
        return "debit"
    if field_name == "credit_amount" or ("贷" in header_text and "借" not in header_text):
        return "credit"
    return "unknown"


def extract_row_summary(row_values: list[Any], headers: list[str]) -> str:
    for col_idx, header in enumerate(headers):
        if match_header(str(header).strip()) == "summary" and col_idx < len(row_values):
            value = row_values[col_idx]
            if value is not None and str(value).strip() and str(value).strip().lower() != "nan":
                return str(value).strip()
    return ""


def signed_amount_from_cell(
    value: Any,
    color: ColorHint,
    *,
    side: AmountSide = "unknown",
    summary: str = "",
) -> Decimal | None:
    """
    按列侧 + 蓝字/红字规则生成带符号发生额（可正可负）。

    - 负数单元格：红字冲减（借/贷列均可）
    - 蓝字 + 正数：同侧正数
    - 红字 + 正数：借/贷列均可能，同侧负数冲减
    - 贷方列红色字体且无冲销摘要：视为正常贷方蓝字发生额（部分软件仅用红色显示贷方列）
    """
    numeric = _amount(value)
    if numeric == Decimal("0.00"):
        return None
    if numeric < Decimal("0.00"):
        return numeric

    if color == "blue":
        return abs(numeric)

    if color == "red":
        if side == "credit" and not is_reversal_summary(summary):
            return abs(numeric)
        return -abs(numeric)

    return numeric


def build_amount_color_grid(
    file_path: str | Path,
    *,
    row_indices: list[int],
    column_count: int,
) -> dict[tuple[int, int], ColorHint]:
    """读取 Excel 数据区单元格字体颜色。"""
    path = Path(file_path)
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return {}

    grid: dict[tuple[int, int], ColorHint] = {}
    try:
        workbook = load_workbook(path, data_only=True, read_only=False)
        worksheet = workbook.active
        for excel_row in row_indices:
            for col_idx in range(column_count):
                cell = worksheet.cell(row=excel_row + 1, column=col_idx + 1)
                hint = classify_font_color(cell)
                if hint != "default":
                    grid[(excel_row, col_idx)] = hint
        workbook.close()
    except Exception:
        return {}
    return grid


def infer_debit_credit_from_colored_row(
    *,
    row_values: list[Any],
    headers: list[str],
    excel_row_index: int,
    color_grid: dict[tuple[int, int], ColorHint],
    has_debit_column: bool,
    has_credit_column: bool,
) -> tuple[Decimal, Decimal, bool]:
    """
    按列方向 + 蓝字/红字规则推断借贷金额（允许负数，同侧冲减）。

    凭证级平衡由后续 _validate_voucher_balance 对同组分录借贷合计校验。
    """
    debit = Decimal("0.00")
    credit = Decimal("0.00")
    used_color = False
    zero = Decimal("0.00")
    summary = extract_row_summary(row_values, headers)

    for col_idx, header in enumerate(headers):
        if col_idx >= len(row_values):
            break
        field_name = match_header(str(header).strip())
        if not is_amount_column(str(header), field_name):
            continue

        color = color_grid.get((excel_row_index, col_idx), "default")
        side = column_amount_side(str(header), field_name)
        signed = signed_amount_from_cell(
            row_values[col_idx],
            color,
            side=side,
            summary=summary,
        )
        if signed is None:
            continue
        if color in {"blue", "red"}:
            used_color = True

        if side == "debit":
            debit += signed
        elif side == "credit":
            credit += signed
        elif has_debit_column and not has_credit_column:
            debit += signed
        elif has_credit_column and not has_debit_column:
            credit += signed
        else:
            if signed > zero:
                debit += signed
            elif signed < zero:
                credit += abs(signed)

    return debit, credit, used_color
