from datetime import date
from decimal import Decimal
from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.db.models import AccountingEntry, OpeningBalance
from app.db.session import Base, get_db
from app.main import app
from app.models.team import Team
from app.models.user import User
from app.services.accounting.export_filename_service import build_report_export_filename
from app.services.accounting.report_export_service import (
    balance_sheet_to_xlsx,
    income_statement_to_xlsx,
    trial_balance_to_xlsx,
)
from app.services.shared.ledger_management_service import create_ledger
from app.db.models import AccountingPeriod


@pytest.fixture
def client():
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(bind=engine)
    TestingSessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)

    def override_get_db():
        db = TestingSessionLocal()
        try:
            yield db
        finally:
            db.close()

    app.dependency_overrides[get_db] = override_get_db
    try:
        with TestClient(app) as test_client:
            yield test_client, TestingSessionLocal
    finally:
        app.dependency_overrides.clear()
        Base.metadata.drop_all(bind=engine)


def _seed(TestingSessionLocal):
    db = TestingSessionLocal()
    try:
        team = Team(name="报表导出团队", type="virtual")
        db.add(team)
        db.flush()
        db.add(User(username="export_user", phone="13800000088", team_id=team.id))
        db.flush()
        ledger = create_ledger(db, team_id=team.id, name="报表导出账簿", accounting_start_date=date(2026, 1, 1))
        db.flush()
        period = db.query(AccountingPeriod).filter(AccountingPeriod.ledger_id == ledger.id).first()
        org_id = period.organization_id
        db.add(OpeningBalance(
            organization_id=org_id, ledger_id=ledger.id, period_id=period.id,
            account_code="1002", debit_balance=Decimal("1000"), credit_balance=Decimal("0"),
        ))
        db.add(OpeningBalance(
            organization_id=org_id, ledger_id=ledger.id, period_id=period.id,
            account_code="4001", debit_balance=Decimal("0"), credit_balance=Decimal("1000"),
        ))
        for code, debit, credit in [
            ("6001", Decimal("0"), Decimal("1000")),
            ("6601", Decimal("100"), Decimal("0")),
            ("1002", Decimal("0"), Decimal("100")),
            ("6801", Decimal("50"), Decimal("0")),
            ("2221", Decimal("0"), Decimal("50")),
        ]:
            db.add(AccountingEntry(
                ledger_id=ledger.id, organization_id=org_id, import_job_id=0,
                voucher_no="导-001", voucher_date=date(2026, 1, 10),
                account_code=code, account_name=code,
                debit_amount=debit, credit_amount=credit, entry_line_no=1,
            ))
        db.commit()
        return ledger.id, period.id
    finally:
        db.close()


def test_build_report_export_filename_pattern():
    name = build_report_export_filename(
        "trial_balance",
        ledger_name="测试账簿",
        period_code="2026-01",
        fmt="xlsx",
        exported_at=__import__("datetime").datetime(2026, 7, 10, 19, 30, 5, tzinfo=__import__("datetime").timezone.utc),
    )
    assert name == "测试账簿_2026-01_科目余额表_20260710_193005.xlsx"


def test_trial_balance_export_xlsx_api(client):
    test_client, TestingSessionLocal = client
    ledger_id, period_id = _seed(TestingSessionLocal)
    test_client.post(f"/api/accounting-periods/{period_id}/pl-transfer")
    resp = test_client.get(
        "/api/reports/trial-balance/export",
        params={"ledger_id": ledger_id, "period_id": period_id, "format": "xlsx"},
    )
    assert resp.status_code == 200
    wb = load_workbook(BytesIO(resp.content), read_only=True)
    assert "科目余额表" in wb.sheetnames
    wb.close()


def test_balance_sheet_export_csv_api(client):
    test_client, TestingSessionLocal = client
    ledger_id, period_id = _seed(TestingSessionLocal)
    test_client.post(f"/api/accounting-periods/{period_id}/pl-transfer")
    resp = test_client.get(
        "/api/reports/balance-sheet/export",
        params={"ledger_id": ledger_id, "period_id": period_id, "format": "csv"},
    )
    assert resp.status_code == 200
    text = resp.content.decode("utf-8-sig")
    assert "资产负债表" in text
    assert "编制单位" in text
    assert "资产总计" in text
    assert "制表人" in text


def test_income_statement_export_unknown_format_400(client):
    test_client, TestingSessionLocal = client
    _, period_id = _seed(TestingSessionLocal)
    resp = test_client.get(
        "/api/reports/income-statement/export",
        params={"period_id": period_id, "format": "doc"},
    )
    assert resp.status_code == 400


def test_trial_balance_export_pdf(client):
    test_client, TestingSessionLocal = client
    ledger_id, period_id = _seed(TestingSessionLocal)
    test_client.post(f"/api/accounting-periods/{period_id}/pl-transfer")
    resp = test_client.get(
        "/api/reports/trial-balance/export",
        params={
            "ledger_id": ledger_id,
            "period_id": period_id,
            "format": "pdf",
            "preparer_name": "张会计",
            "reviewer_name": "李复核",
            "approver_name": "王主管",
        },
    )
    assert resp.status_code == 200
    assert resp.headers["content-type"] == "application/pdf"
    assert resp.content[:4] == b"%PDF"


def test_cash_flow_export_xlsx_api(client):
    test_client, TestingSessionLocal = client
    ledger_id, period_id = _seed(TestingSessionLocal)
    resp = test_client.get(
        "/api/reports/cash-flow-statement/export",
        params={"ledger_id": ledger_id, "period_id": period_id, "format": "xlsx"},
    )
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]


def test_reports_package_export_zip(client):
    test_client, TestingSessionLocal = client
    ledger_id, period_id = _seed(TestingSessionLocal)
    test_client.post(f"/api/accounting-periods/{period_id}/pl-transfer")
    resp = test_client.get(
        "/api/reports/package/export",
        params={"ledger_id": ledger_id, "period_id": period_id, "include_pdf": "true"},
    )
    assert resp.status_code == 200
    assert resp.headers["content-type"] == "application/zip"
    import zipfile
    from io import BytesIO
    with zipfile.ZipFile(BytesIO(resp.content)) as zf:
        names = zf.namelist()
    assert "01_科目余额表.xlsx" in names
    assert "02_资产负债表.xlsx" in names
    assert "03_损益表.xlsx" in names
    assert "04_现金流量表.xlsx" in names
    assert "01_科目余额表_签章版.pdf" in names


def test_report_service_builders_minimal():
    trial = trial_balance_to_xlsx({
        "rows": [{"account_code": "1002", "account_name": "银行", "category": "asset",
                  "opening_debit": "0", "opening_credit": "0", "period_debit": "0",
                  "period_credit": "0", "closing_debit": "100", "closing_credit": "0"}],
        "totals": {"opening_debit": "0", "opening_credit": "0", "period_debit": "0",
                   "period_credit": "0", "closing_debit": "100", "closing_credit": "100"},
    })
    assert len(trial) > 100
    bs = balance_sheet_to_xlsx({"assets": [], "liabilities": [], "equity": [],
                                "assets_total": "0", "liabilities_total": "0", "equity_total": "0",
                                "is_balanced": True})
    assert len(bs) > 100
    inc = income_statement_to_xlsx({"revenue": {}, "expense": {}, "operating_profit": "0",
                                    "total_profit": "0", "income_tax": "0", "net_profit": "0"})
    assert len(inc) > 100
