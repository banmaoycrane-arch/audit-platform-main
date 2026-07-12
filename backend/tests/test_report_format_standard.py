"""报表标准格式单元测试。"""
from openpyxl import load_workbook
import io

from app.services.accounting.report_export_service import trial_balance_to_xlsx
from app.services.accounting.report_format_standard import TRIAL_BALANCE_HEADERS, category_label


def test_category_label_cn():
    assert category_label("asset") == "资产"
    assert category_label("equity") == "所有者权益"


def test_trial_balance_xlsx_has_standard_headers():
    report = {
        "period_code": "2026-01",
        "ledger_name": "测试账簿",
        "rows": [
            {
                "account_code": "1001",
                "account_name": "库存现金",
                "category": "asset",
                "opening_debit": 100,
                "opening_credit": 0,
                "period_debit": 50,
                "period_credit": 20,
                "ytd_debit": 50,
                "ytd_credit": 20,
                "closing_debit": 130,
                "closing_credit": 0,
            }
        ],
        "totals": {
            "opening_debit": 100,
            "opening_credit": 0,
            "period_debit": 50,
            "period_credit": 20,
            "ytd_debit": 50,
            "ytd_credit": 20,
            "closing_debit": 130,
            "closing_credit": 0,
        },
    }
    body = trial_balance_to_xlsx(report, ledger_name="测试账簿")
    wb = load_workbook(io.BytesIO(body))
    ws = wb.active
    header_row = None
    for row in ws.iter_rows(values_only=True):
        if row and row[0] == TRIAL_BALANCE_HEADERS[0]:
            header_row = row
            break
    assert header_row is not None
    assert header_row[2] == "科目类别"
