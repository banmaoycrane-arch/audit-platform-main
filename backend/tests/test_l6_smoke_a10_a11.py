# -*- coding: utf-8 -*-
"""
L6 路径 A 冒烟测试：A10 期末处理 + A11 财务报表与导出

运行方式：
    cd backend
    pytest tests/test_l6_smoke_a10_a11.py -v

对应验收步骤：
    A10 — 损益结转、结账、期间状态流转
    A11 — 科目余额表 / 资产负债表 / 利润表 JSON + xlsx 导出
"""
from datetime import date
from decimal import Decimal

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.db.models import AccountingEntry, AccountingPeriod, OpeningBalance
from app.db.session import Base, get_db
from app.main import app
from app.models.team import Team
from app.models.user import User
from app.services.shared.ledger_management_service import create_ledger


@pytest.fixture
def client():
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(bind=engine)
    TestingSessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)

    def override_get_db():
        db = TestingSessionLocal()
        try:
            yield db
        finally:
            db.close()

    app.dependency_overrides[get_db] = override_get_db
    try:
        with TestClient(app) as test_client:
            yield test_client, TestingSessionLocal
    finally:
        app.dependency_overrides.clear()
        Base.metadata.drop_all(bind=engine)


def _seed_smoke_ledger(TestingSessionLocal):
    """
    标准样例：期初 1002/4001 各 1000；本期收入 1000、费用 100、所得税 50 → 净利润 850。
    与 test_period_close_pl_transfer 一致，便于 A10/A11 回归。
    """
    db = TestingSessionLocal()
    try:
        team = Team(name="L6冒烟团队", type="virtual")
        db.add(team)
        db.flush()

        user = User(username="l6_smoke_user", phone="13800000099", team_id=team.id)
        db.add(user)
        db.flush()

        ledger = create_ledger(
            db,
            team_id=team.id,
            name="L6冒烟账簿",
            accounting_start_date=date(2026, 1, 1),
        )
        db.flush()

        period = db.query(AccountingPeriod).filter(
            AccountingPeriod.ledger_id == ledger.id,
            AccountingPeriod.period_code == "2026-01",
        ).first()
        org_id = period.organization_id

        db.add(OpeningBalance(
            organization_id=org_id,
            ledger_id=ledger.id,
            period_id=period.id,
            account_code="1002",
            debit_balance=Decimal("1000"),
            credit_balance=Decimal("0"),
        ))
        db.add(OpeningBalance(
            organization_id=org_id,
            ledger_id=ledger.id,
            period_id=period.id,
            account_code="4001",
            debit_balance=Decimal("0"),
            credit_balance=Decimal("1000"),
        ))

        for code, debit, credit in [
            ("1122", Decimal("1130"), Decimal("0")),
            ("6001", Decimal("0"), Decimal("1000")),
            ("2221", Decimal("0"), Decimal("130")),
            ("6601", Decimal("100"), Decimal("0")),
            ("1002", Decimal("0"), Decimal("100")),
            ("6801", Decimal("50"), Decimal("0")),
            ("2221", Decimal("0"), Decimal("50")),
        ]:
            db.add(AccountingEntry(
                ledger_id=ledger.id,
                organization_id=org_id,
                import_job_id=0,
                voucher_no="L6-001",
                voucher_date=date(2026, 1, 15),
                account_code=code,
                account_name=code,
                debit_amount=debit,
                credit_amount=credit,
                entry_line_no=1,
                post_status="posted",
            ))
        db.commit()
        return {
            "ledger_id": ledger.id,
            "org_id": org_id,
            "period_id": period.id,
            "period_code": period.period_code,
        }
    finally:
        db.close()


def _assert_xlsx_has_sheet(content: bytes, sheet_name: str):
    wb = load_workbook(filename=__import__("io").BytesIO(content), read_only=True)
    assert sheet_name in wb.sheetnames
    wb.close()


class TestL6SmokeA10PeriodClose:
    """A10：期末处理冒烟。"""

    def test_a10_pl_transfer_then_close(self, client):
        test_client, TestingSessionLocal = client
        ctx = _seed_smoke_ledger(TestingSessionLocal)
        period_id = ctx["period_id"]
        org_id = ctx["org_id"]

        # 结转前资产负债表不平衡
        bs_before = test_client.get(
            "/api/reports/balance-sheet",
            params={"organization_id": org_id, "period_id": period_id},
        )
        assert bs_before.status_code == 200
        assert bs_before.json()["is_balanced"] is False

        # 未结转不能结账
        close_blocked = test_client.post(f"/api/accounting-periods/{period_id}/close")
        assert close_blocked.status_code == 400
        assert "尚未结转损益" in close_blocked.json()["detail"]

        # 损益结转
        pl = test_client.post(f"/api/accounting-periods/{period_id}/pl-transfer")
        assert pl.status_code == 200
        pl_body = pl.json()
        assert pl_body["status"] == "pl_transferred"
        assert pl_body["net_profit"] in (850, "850", "850.00", 850.0)

        # 结转后资产负债表平衡
        bs_after = test_client.get(
            "/api/reports/balance-sheet",
            params={"organization_id": org_id, "period_id": period_id},
        )
        assert bs_after.status_code == 200
        assert bs_after.json()["is_balanced"] is True

        # 结账
        close_ok = test_client.post(
            f"/api/accounting-periods/{period_id}/close",
            json={"operator": "l6_smoke", "reason": "A10冒烟结账"},
        )
        assert close_ok.status_code == 200
        assert close_ok.json()["status"] == "closed"


class TestL6SmokeA11Reports:
    """A11：财务报表 JSON + 导出冒烟。"""

    def test_a11_reports_after_pl_transfer(self, client):
        test_client, TestingSessionLocal = client
        ctx = _seed_smoke_ledger(TestingSessionLocal)
        period_id = ctx["period_id"]
        ledger_id = ctx["ledger_id"]
        org_id = ctx["org_id"]

        test_client.post(f"/api/accounting-periods/{period_id}/pl-transfer")

        trial = test_client.get(
            "/api/reports/trial-balance",
            params={"ledger_id": ledger_id, "period_id": period_id},
        )
        assert trial.status_code == 200
        assert trial.json()["is_balanced"] is True

        balance = test_client.get(
            "/api/reports/balance-sheet",
            params={"ledger_id": ledger_id, "period_id": period_id},
        )
        assert balance.status_code == 200
        assert balance.json()["is_balanced"] is True

        income = test_client.get(
            "/api/reports/income-statement",
            params={"organization_id": org_id, "period_id": period_id},
        )
        assert income.status_code == 200
        assert Decimal(str(income.json()["net_profit"])) == Decimal("850")

    def test_a11_report_xlsx_exports(self, client):
        test_client, TestingSessionLocal = client
        ctx = _seed_smoke_ledger(TestingSessionLocal)
        period_id = ctx["period_id"]
        ledger_id = ctx["ledger_id"]

        test_client.post(f"/api/accounting-periods/{period_id}/pl-transfer")

        exports = [
            ("/api/reports/trial-balance/export", "科目余额表"),
            ("/api/reports/balance-sheet/export", "资产负债表"),
            ("/api/reports/income-statement/export", "利润表"),
            ("/api/reports/cash-flow-statement/export", "现金流量表"),
        ]
        for path, sheet_name in exports:
            resp = test_client.get(path, params={"ledger_id": ledger_id, "period_id": period_id, "format": "xlsx"})
            assert resp.status_code == 200, f"{path} failed: {resp.text}"
            assert "spreadsheetml" in resp.headers["content-type"]
            assert "attachment" in resp.headers["content-disposition"]
            assert len(resp.content) > 200
            _assert_xlsx_has_sheet(resp.content, sheet_name)


class TestL6SmokeFullChain:
    """A10 → A11 串联冒烟（发布前快速回归）。"""

    def test_l6_smoke_full_chain(self, client):
        test_client, TestingSessionLocal = client
        ctx = _seed_smoke_ledger(TestingSessionLocal)
        period_id = ctx["period_id"]
        ledger_id = ctx["ledger_id"]

        # A10
        pl = test_client.post(f"/api/accounting-periods/{period_id}/pl-transfer")
        assert pl.status_code == 200

        close = test_client.post(
            f"/api/accounting-periods/{period_id}/close",
            json={"operator": "l6_smoke", "reason": "full chain"},
        )
        assert close.status_code == 200
        assert close.json()["status"] == "closed"

        # A11 — 结账后仍可取报表（快照/即时口径由服务决定，此处验证接口可用）
        for path in [
            "/api/reports/trial-balance",
            "/api/reports/balance-sheet",
            "/api/reports/income-statement",
        ]:
            resp = test_client.get(path, params={"ledger_id": ledger_id, "period_id": period_id})
            assert resp.status_code == 200

        export = test_client.get(
            "/api/reports/trial-balance/export",
            params={"ledger_id": ledger_id, "period_id": period_id, "format": "xlsx"},
        )
        assert export.status_code == 200
        assert len(export.content) > 200

        package = test_client.get(
            "/api/reports/package/export",
            params={"ledger_id": ledger_id, "period_id": period_id, "include_pdf": "true"},
        )
        assert package.status_code == 200
        assert package.headers["content-type"] == "application/zip"
