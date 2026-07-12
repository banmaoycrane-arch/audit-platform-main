"""同一凭证内重复借方分录不应被去重误删。"""

from __future__ import annotations

from datetime import date
from decimal import Decimal
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font

from app.services.audit.audit_day_book_service import _build_day_book_report, _entry_duplicate_key
from app.services.doc_parsing.file_parser_service import parse_structured_accounting_entries


def test_identical_debit_lines_within_voucher_are_not_deduped() -> None:
    """记-0011 场景：4 笔相同借方 + 4 笔不同贷方，借方合计应为 8 万。"""
    base = {
        "voucher_no": "记-0011",
        "voucher_date": date(2022, 7, 26),
        "summary": "2022-07-26, 小额普通贷记来账",
        "account_code": "100202",
        "account_name": "银行存款_农商行",
        "debit_amount": Decimal("20000.00"),
        "credit_amount": Decimal("0.00"),
        "parse_group_key": "记-0011|2022-07-26|g1",
    }
    entries: list[dict] = []
    for line_no in range(1, 5):
        row = dict(base)
        row["entry_line_no"] = line_no
        entries.append(row)

    credit_accounts = [
        ("2241", "其他应付款_0372 唐山鑫虎重型矿山机械"),
        ("2202", "应付账款_0044 沈阳盛世五寰科技有限公司"),
        ("2241", "其他应付款_0373 成都利君实业股份"),
        ("2241", "其他应付款_0374 马鞍山格林环保科技"),
    ]
    for line_no, (code, name) in enumerate(credit_accounts, start=5):
        entries.append(
            {
                "voucher_no": "记-0011",
                "voucher_date": date(2022, 7, 26),
                "summary": "2022-07-26, 小额普通贷记来账",
                "account_code": code,
                "account_name": name,
                "debit_amount": Decimal("0.00"),
                "credit_amount": Decimal("20000.00"),
                "parse_group_key": "记-0011|2022-07-26|g1",
                "entry_line_no": line_no,
            }
        )

    keys = {_entry_duplicate_key(entry) for entry in entries}
    assert len(keys) == 8

    seen: set[tuple[str, ...]] = set()
    kept: list[dict] = []
    for entry in entries:
        key = _entry_duplicate_key(entry)
        if key in seen:
            continue
        seen.add(key)
        kept.append(entry)
    assert len(kept) == 8

    report = _build_day_book_report(kept)
    assert report.unbalanced_count == 0
    item = report.unbalanced_vouchers[0] if report.unbalanced_vouchers else None
    assert item is None


def test_voucher_0011_like_excel_parses_balanced(tmp_path: Path) -> None:
    path = tmp_path / "voucher_0011.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    headers = ["凭证日期", "凭证号", "摘要", "科目", "借方金额", "贷方金额", "制单人"]
    worksheet.append(headers)
    rows = [
        ["2022-07-26", "记-0011", "2022-07-22, 小额普通贷记来账", "100202 银行存款_农商行", 20000, None, "吴林玉"],
        ["", "", "2022-07-25, 小额普通贷记来账", "100202 银行存款_农商行", 20000, None, ""],
        ["", "", "2022-07-26, 小额普通贷记来账", "100202 银行存款_农商行", 20000, None, ""],
        ["", "", "2022-07-26, 小额普通贷记来账", "100202 银行存款_农商行", 20000, None, ""],
        ["", "", "2022-07-22, 小额普通贷记来账", "2241 其他应付款_0372 唐山鑫虎", None, 20000, ""],
        ["", "", "2022-07-25, 小额普通贷记来账", "2202 应付账款_0044 沈阳盛世", None, 20000, ""],
        ["", "", "2022-07-26, 小额普通贷记来账", "2241 其他应付款_0373 成都利君", None, 20000, ""],
        ["", "", "2022-07-26, 小额普通贷记来账", "2241 其他应付款_0374 马鞍山格林", None, 20000, ""],
    ]
    for row in rows:
        worksheet.append(row)
        excel_row = worksheet.max_row
        if row[4]:
            worksheet.cell(excel_row, 5).font = Font(color="0000FF")
        if row[5]:
            worksheet.cell(excel_row, 6).font = Font(color="FF0000")
    workbook.save(path)

    result = parse_structured_accounting_entries(str(path))
    assert len(result.entries) == 8
    debit_total = sum(entry["debit_amount"] for entry in result.entries)
    credit_total = sum(entry["credit_amount"] for entry in result.entries)
    assert debit_total == Decimal("80000.00")
    assert credit_total == Decimal("80000.00")

    report = _build_day_book_report(result.entries)
    assert report.unbalanced_count == 0
